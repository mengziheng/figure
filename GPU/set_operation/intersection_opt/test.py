import subprocess
import openpyxl
import re
import os
from openpyxl.styles import Alignment

# 定义要执行的程序
program_path = "your_program.exe"

# 定义y值
# y_values = [32, 64]
y_values = [32, 64, 128, 256, 512, 1024, 2048]

# 执行程序并记录结果到Excel文件中
def execute_and_record():
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(row=1, column=1, value="hashTable/array")
    
    # 写入第一行，记录y值
    for i, y in enumerate(y_values, start=1):
        sheet.cell(row=1, column=i+1, value=y)

    # 写入第一列，记录x值
    for i, x in enumerate(y_values, start=1):
        sheet.cell(row=i+1, column=1, value=x)

    # 执行程序并记录结果
    for i, y in enumerate(y_values):
        for j, x in enumerate(y_values):
            if x <= y:
                result = []
                command = f"./ncu_run.sh {x} {y}"
                regex_pattern = r"sector\s+(\d+,*\d*)"
                process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)

                for line in iter(process.stdout.readline, b''):
                    line = line.decode().strip()
                    match = re.search(regex_pattern, line)
                    if match:
                        result.append(str(match.group(1)))

                align = Alignment(horizontal='right')
                sheet.cell(row=i+2, column=j+2, value=f"{result[0]}/{result[1]}").alignment = align

    # 保存Excel文件
    wb.save("results.xlsx")

if __name__ == "__main__":
    command = f"nvcc -o test test_set.cu"
    os.system(command)
    execute_and_record()
